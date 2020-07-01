from django.core.management.base import BaseCommand, CommandError

from market.models import Category, Product

from prj.settings import DATA_DIR

from openpyxl import load_workbook


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('Clearing DB')
        Category.objects.all().delete()
        Product.objects.all().delete()

        print('Start importing excel {}'.format(DATA_DIR))
        wb = load_workbook(DATA_DIR+'/price.xlsx')
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        cat = None

        for cnt in range(1, sheet.max_row+1):
            item = sheet.cell(row=cnt, column=3).value
            id = sheet.cell(row=cnt, column=2).value
            if id == None:
                print('Create new category')
                cat = Category()
                cat.name = item
                cat.save()
            else:
                print('Create new food')
                if cat:
                    p = Product()
                    p.name = item
                    p.category = cat
                    p.save()


